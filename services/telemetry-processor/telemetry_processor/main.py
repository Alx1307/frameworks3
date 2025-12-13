import csv
import time
import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .generator import TelemetryGenerator
from .database import DatabaseManager
from .config import app_config

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class CSVExporter:
    def __init__(self, output_dir: str = None):
        self.output_dir = Path(output_dir or app_config.csv_output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_csv(self, records) -> Path:
        timestamp = datetime.now()
        filename = f"telemetry_{timestamp.strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = self.output_dir / filename
        
        fieldnames = [
            'timestamp',
            'is_operational',
            'voltage',
            'temperature',
            'sensor_id',
            'source_file'
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for record in records:
                writer.writerow(record.to_csv_dict())
        
        logger.info(f"CSV file created: {filepath}")
        return filepath


class ExcelExporter:
    def __init__(self, output_dir: str = None):
        self.output_dir = Path(output_dir or app_config.excel_output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(self, records) -> Path:
        timestamp = datetime.now()
        filename = f"telemetry_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Telemetry Data"
        
        headers = [
            "Timestamp", 
            "Operational", 
            "Voltage (V)", 
            "Temperature (°C)", 
            "Sensor ID", 
            "Source File"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", 
                end_color="366092", 
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        for row_idx, record in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, 
                   value=record.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_idx, column=2, 
                   value="TRUE" if record.is_operational else "FALSE")
            ws.cell(row=row_idx, column=3, value=record.voltage)
            ws.cell(row=row_idx, column=4, value=record.temperature)
            ws.cell(row=row_idx, column=5, value=record.sensor_id)
            ws.cell(row=row_idx, column=6, value=record.source_file)
            
            status_cell = ws.cell(row=row_idx, column=2)
            if record.is_operational:
                status_cell.font = Font(color="00AA00", bold=True)
            else:
                status_cell.font = Font(color="FF0000", bold=True)
        
        self._add_summary_sheet(wb, records, timestamp)
        
        wb.save(filepath)
        logger.info(f"Excel file created: {filepath}")
        return filepath
    
    def _add_summary_sheet(self, wb, records, timestamp):
        ws_summary = wb.create_sheet(title="Summary")
        
        ws_summary.cell(row=1, column=1, value="Report Summary").font = Font(bold=True, size=14)
        
        summary_data = [
            ("Report Generated", timestamp.strftime('%Y-%m-%d %H:%M:%S')),
            ("Total Records", len(records)),
            ("Operational Systems", sum(1 for r in records if r.is_operational)),
            ("Average Voltage", f"{sum(r.voltage for r in records)/len(records):.2f} V"),
            ("Average Temperature", f"{sum(r.temperature for r in records)/len(records):.2f} °C"),
        ]
        
        for i, (label, value) in enumerate(summary_data, 3):
            ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=i, column=2, value=value)


class TelemetryProcessor:
    def __init__(self):
        self.generator = TelemetryGenerator(records_per_batch=5)
        self.csv_exporter = CSVExporter()
        self.excel_exporter = ExcelExporter()
        self.db_manager = DatabaseManager()
    
    def run_cycle(self):
        logger.info("Starting telemetry processing cycle")
        
        try:
            records = self.generator.generate_batch()
            logger.info(f"Generated {len(records)} telemetry records")
            
            csv_file = self.csv_exporter.export_to_csv(records)
            
            excel_file = self.excel_exporter.export_to_excel(records)
            
            self.db_manager.save_records(records)
            
            logger.info(f"Cycle completed successfully. Files: {csv_file.name}, {excel_file.name}")
            
        except Exception as e:
            logger.error(f"Error in processing cycle: {e}")
            raise
    
    def run_continuously(self):
        logger.info(f"Starting telemetry processor with {app_config.generation_interval} second interval")
        
        while True:
            try:
                self.run_cycle()
            except Exception as e:
                logger.error(f"Critical error: {e}")
            
            time.sleep(app_config.generation_interval)


def main():
    processor = TelemetryProcessor()
    processor.run_continuously()


if __name__ == "__main__":
    main()